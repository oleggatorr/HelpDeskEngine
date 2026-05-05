import re
import copy
from typing import Dict, Any, List, Optional, Tuple
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

class ExcelTemplateEngine:
    """
    Движок генерации Excel по шаблонам с поддержкой:
        - Переменных: {{field}}
        - Условий: {{@if:field}} ... {{@else}} ... {{@endif}}
        - Циклов: {{@each:field}} ... {{@endeach}}
    Все стили, форматы и базовые объединения ячеек сохраняются.
    """
    def __init__(self):
        self._dir_if = re.compile(r'\{\{@if:(.+?)\}\}')
        self._dir_else = re.compile(r'\{\{@else\}\}')
        self._dir_endif = re.compile(r'\{\{@endif\}\}')
        self._dir_each = re.compile(r'\{\{@each:(.+?)\}\}')
        self._dir_endeach = re.compile(r'\{\{@endeach\}\}')
        self._placeholder = re.compile(r'\{\{(.+?)\}\}')

    def render(self, template_path: str, data: Dict[str, Any], output_path: str) -> str:
        wb = load_workbook(template_path)
        for ws in wb.worksheets:
            self._process_sheet(ws, data)
        wb.save(output_path)
        return output_path

    def _process_sheet(self, ws, data: Dict[str, Any]) -> None:
        blocks = self._parse_to_blocks(ws)
        self._render_blocks(blocks, data, ws)

    def _parse_to_blocks(self, ws) -> List[Dict]:
        """Превращает строки листа в AST-подобную структуру блоков."""
        blocks = []
        stack: List[Dict] = []  # Стек для вложенных if/each

        for row_idx in range(1, ws.max_row + 1):
            row_vals = [c.value for c in ws[row_idx]]
            directive = self._find_directive(row_vals)
            
            if not directive:
                block = {'type': 'row', 'idx': row_idx, 'vals': row_vals}
                self._push_to_stack(stack, block)
                continue

            dtype, payload = directive
            if dtype == 'if':
                block = {'type': 'if', 'condition': payload, 'children': [], 'else_children': [], 'is_else': False}
                stack.append(block)
            elif dtype == 'else':
                if stack and stack[-1]['type'] == 'if':
                    stack[-1]['is_else'] = True
            elif dtype == 'endif':
                if stack and stack[-1]['type'] == 'if':
                    finished = stack.pop()
                    self._push_to_stack(stack, finished)
            elif dtype == 'each':
                block = {'type': 'each', 'var': payload, 'children': []}
                stack.append(block)
            elif dtype == 'endeach':
                if stack and stack[-1]['type'] == 'each':
                    finished = stack.pop()
                    self._push_to_stack(stack, finished)

        # Выгружаем оставшиеся блоки из стека (если пользователь забыл закрыть)
        while stack:
            self._push_to_stack([], stack.pop())
        return blocks

    def _push_to_stack(self, stack: List, block: Dict) -> None:
        if not stack:
            self._global_blocks.append(block) if hasattr(self, '_global_blocks') else None
        else:
            parent = stack[-1]
            target = parent.get('else_children') if parent.get('is_else') else parent['children']
            target.append(block)

    def _render_blocks(self, blocks: List[Dict], data: Dict, target_ws) -> None:
        """Рендерит блоки в целевой лист, сохраняя стили."""
        # Создаём временный лист для рендера, затем заменяем исходный
        temp_ws = target_ws.parent.create_sheet(f"__temp_{target_ws.title}")
        target_ws.parent.remove(target_ws)
        temp_ws.title = target_ws.title

        row_counter = 1
        for block in blocks:
            row_counter = self._execute_block(block, data, temp_ws, row_counter)
        
        # Копируем merged_cells
        self._restore_merges(temp_ws, target_ws)

    def _execute_block(self, block: Dict, data: Dict, ws, start_row: int) -> int:
        if block['type'] == 'row':
            return self._render_row(block['idx'], data, ws, start_row)
        elif block['type'] == 'if':
            cond_val = self._resolve_path(block['condition'], data)
            if cond_val:
                for child in block['children']:
                    start_row = self._execute_block(child, data, ws, start_row)
            else:
                for child in block['else_children']:
                    start_row = self._execute_block(child, data, ws, start_row)
            return start_row
        elif block['type'] == 'each':
            items = self._resolve_path(block['var'], data)
            if not isinstance(items, list):
                items = [items] if items is not None else []
            for item in items:
                for child in block['children']:
                    start_row = self._execute_block(child, data | {block['var'].split('.')[-1]: item}, ws, start_row)
            return start_row
        return start_row

    def _render_row(self, src_row_idx: int, data: Dict, dst_ws, dst_row: int) -> int:
        src_ws = self._source_ws
        for col in range(1, src_ws.max_column + 1):
            src_cell = src_ws.cell(row=src_row_idx, column=col)
            dst_cell = dst_ws.cell(row=dst_row, column=col)
            
            # Копируем значение с подстановкой плейсхолдеров
            val = src_cell.value
            if isinstance(val, str):
                # Убираем директивы
                val = self._dir_if.sub('', val)
                val = self._dir_else.sub('', val)
                val = self._dir_endif.sub('', val)
                val = self._dir_each.sub('', val)
                val = self._dir_endeach.sub('', val)
                val = self._placeholder.sub(lambda m: str(self._resolve_path(m.group(1), data)), val)
            dst_cell.value = val
            self._copy_style(src_cell, dst_cell)
        return dst_row + 1

    def _copy_style(self, src, dst) -> None:
        if src.has_style:
            dst.font = src.font.copy()
            dst.fill = src.fill.copy()
            dst.alignment = src.alignment.copy()
            dst.border = src.border.copy()
            dst.number_format = src.number_format
            dst.protection = src.protection.copy()

    def _resolve_path(self, path: str, data: Dict) -> Any:
        parts = path.strip().split('.')
        val = data
        for p in parts:
            if isinstance(val, dict):
                val = val.get(p)
            else:
                return None
        return val

    def _find_directive(self, row_vals: List) -> Optional[Tuple[str, str]]:
        for v in row_vals:
            if not isinstance(v, str): continue
            for pat, dtype in [
                (self._dir_if, 'if'), (self._dir_else, 'else'), (self._dir_endif, 'endif'),
                (self._dir_each, 'each'), (self._dir_endeach, 'endeach')
            ]:
                m = pat.search(v)
                if m: return dtype, m.group(1)
        return None

    def _restore_merges(self, src, dst) -> None:
        for mc in src.merged_cells.ranges:
            dst.merge_cells(start_row=mc.min_row, start_column=mc.min_col,
                            end_row=mc.max_row, end_column=mc.max_col)