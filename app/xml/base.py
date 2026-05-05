import os
from typing import Any, List, Optional, Dict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.exceptions import InvalidFileException

class ExcelManager:
    """
    OOP-обёртка для работы с .xlsx: чтение, запись, стили, объединение.
    Индексы строк/столбцов: 1-based (как в Excel).
    """
    def __init__(self, filepath: str, sheet_name: str = "Sheet1", create_if_missing: bool = False):
        self.filepath = filepath
        self.sheet_name = sheet_name
        self.create_if_missing = create_if_missing
        self.wb: Optional[Workbook] = None
        self.ws = None
        self._load_or_create()

    def _load_or_create(self) -> None:
        if os.path.exists(self.filepath):
            try:
                self.wb = load_workbook(self.filepath)
                if self.sheet_name in self.wb.sheetnames:
                    self.ws = self.wb[self.sheet_name]
                else:
                    self.ws = self.wb.create_sheet(self.sheet_name)
            except InvalidFileException as e:
                raise ValueError(f"Файл повреждён: {e}")
        else:
            if not self.create_if_missing:
                raise FileNotFoundError(f"Файл '{self.filepath}' не найден.")
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = self.sheet_name

    # ==================== ЗАПИСЬ / ЧТЕНИЕ ====================
    def write_cell(self, row: int, col: int, value: Any) -> None:
        if row < 1 or col < 1:
            raise ValueError("Индексы должны быть >= 1")
        self.ws.cell(row=row, column=col, value=value)

    def write_row(self, row: int, data: List[Any], start_col: int = 1) -> None:
        for i, val in enumerate(data):
            self.write_cell(row, start_col + i, val)

    def append_row(self, data: List[Any]) -> int:
        self.ws.append(data)
        return self.ws.max_row

    def read_cell(self, row: int, col: int) -> Any:
        return self.ws.cell(row=row, column=col).value

    # ==================== СТИЛИ И ФОРМАТИРОВАНИЕ ====================
    def style_cell(self, row: int, col: int,
                   font: Optional[Dict[str, Any]] = None,
                   fill: Optional[Dict[str, Any]] = None,
                   alignment: Optional[Dict[str, Any]] = None,
                   border: Optional[Dict[str, Any]] = None,
                   number_format: Optional[str] = None) -> None:
        """Применить стили к одной ячейке."""
        cell = self.ws.cell(row=row, column=col)

        if font:
            cell.font = Font(**font)
        if fill:
            fg = fill.get("fg_color", "FFFFFF")
            cell.fill = PatternFill(start_color=fg, end_color=fg, fill_type="solid")
        if alignment:
            cell.alignment = Alignment(**alignment)
        if border:
            # Простой border: передаём dict {'left': 'thin', 'top': 'medium', ...}
            sides = {}
            for side_name, style in border.items():
                sides[f"{side_name}_side"] = Side(style=style)
            cell.border = Border(**sides)
        if number_format:
            cell.number_format = number_format

    def style_range(self, start_row: int, start_col: int, 
                    end_row: int, end_col: int, **style_kwargs) -> None:
        """Применить одинаковые стили к диапазону ячеек."""
        for r in range(start_row, end_row + 1):
            for c in range(start_col, end_col + 1):
                self.style_cell(r, c, **style_kwargs)

    def merge_cells(self, start_row: int, start_col: int, end_row: int, end_col: int) -> None:
        """Объединить ячейки в прямоугольную область."""
        if start_row > end_row or start_col > end_col:
            raise ValueError("Некорректный диапазон для объединения.")
        self.ws.merge_cells(start_row=start_row, start_column=start_col,
                            end_row=end_row, end_column=end_col)

    # ==================== СОХРАНЕНИЕ ====================
    def save(self, filepath: Optional[str] = None) -> None:
        target = filepath or self.filepath
        self.wb.save(target)
        if filepath:
            self.filepath = target

    def close(self) -> None:
        if self.wb:
            self.save()
            self.wb = None
            self.ws = None

    def __enter__(self):
        return self
    def __exit__(self, *args):
        self.close()